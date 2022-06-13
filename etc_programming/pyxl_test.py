from openpyxl import Workbook
wb = Workbook()

ws = wb.active # 활성화된 워크시트 선택
ws['A1'] = 42 # 셀에 데이터 추가

ws.append([1, 2, 3]) # 그 다음 행에 데이터 추가

import datetime
ws['A2'] = datetime.datetime.now()

wb.save("sample.xlsx") # 저장